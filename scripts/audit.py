import json
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

folder_path = r'C:\Users\91745\OneDrive\Desktop\stg_hackathon\data\output_js'
xlsx_file_path = r'C:\Users\91745\OneDrive\Desktop\stg_hackathon\utils\Audit_report.xlsx'

yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

headers = [
    "Date & Time", "Machine Name", "Process Name", "Process Start Time", 
    "Process End Time", "", "Candidate Name", "Details (As JSON)", 
    "Status", "Exception", "Remarks"
]

data = []

def append_to_data(candidate_name, json_data, file_name):
    data.append({
        "Date & Time": "",
        "Machine Name": "",
        "Process Name": "",
        "Process Start Time": "",
        "Process End Time": "",
        "": "",
        "Candidate Name": candidate_name,
        "Details (As JSON)": json.dumps(json_data),
        "Status": "",
        "Exception": "",
        "Remarks": ""
    })

for file_name in os.listdir(folder_path):
    file_path = os.path.join(folder_path, file_name)

    if file_name.endswith('.txt'):
        with open(file_path, 'r') as txt_file:
            json_content = txt_file.read().strip()
            try:
                json_data = json.loads(json_content)
                candidate_name = json_data.pop("name", "Unknown")
                append_to_data(candidate_name, json_data, file_name)
            except json.JSONDecodeError as e:
                print(f"Error decoding JSON in file {file_name}: {e}")

df = pd.DataFrame(data, columns=headers)
df.to_excel(xlsx_file_path, sheet_name="Audit Report", index=False)

wb = load_workbook(xlsx_file_path)
ws = wb.active

for cell in ws[1]:
    cell.fill = yellow_fill

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
    for cell in row:
        cell.fill = yellow_fill

wb.save(xlsx_file_path)

print(f"Data with styled headers and column has been successfully saved to {xlsx_file_path}")
